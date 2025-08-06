# Manipulating Excel files.

import openpyxl

# for 1 file.
'''
wb = openpyxl.load_workbook('octobre.xlsx')
print(wb.sheetnames)
sheet = wb[wb.sheetnames[0]]
# print(sheet['A1'].value)

# cell = sheet.cell(row=1, column=2)
# print(cell.value)

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        print(sheet.cell(row=row, column=col).value)
'''

# for multiple files.

wb1 = openpyxl.load_workbook('octobre.xlsx', data_only = True)
wb2 = openpyxl.load_workbook('novembre.xlsx', data_only = True)
wb3 = openpyxl.load_workbook('decembre.xlsx', data_only = True)


def add_datas_wb(wb, d):
    sheet = wb.active
    for row in range(2, sheet.max_row):
        name_article = sheet.cell(row=row, column=1).value
        if not name_article:
            break
        total_sales = sheet.cell(row=row, column=4).value
        if d.get(name_article):
            d[name_article].append(total_sales)
        else:
            d[name_article] = [total_sales]


datas = {}
add_datas_wb(wb1, datas)
add_datas_wb(wb2, datas)
add_datas_wb(wb3, datas)
print(datas)


wb_output = openpyxl.Workbook()
sheet = wb_output.active
sheet['A1'] = 'Article' 
sheet['B1'] = 'Octobre' 
sheet['C1'] = 'Novembre' 
sheet['D1'] = 'Decembre' 

row = 2
for i in datas.items():
    name_article = i[0]
    sales = i[1]
    sheet.cell(row, 1).value = name_article
    for c in range(0, len(sales)):
        sheet.cell(row, 2+c).value = sales[c]
    row += 1


chart_ref = openpyxl.chart.Reference(sheet, min_col = 2, min_row = 2, max_col = sheet.max_column, max_row = 2)
chart_serie = openpyxl.chart.Series(chart_ref, title = 'Sales in Euros')
chart = openpyxl.chart.BarChart()
chart.title = 'Apple Price Evolution'
chart.append(chart_serie)

sheet.add_chart(chart, 'F2')


wb_output.save('output_excel_file.xlsx')